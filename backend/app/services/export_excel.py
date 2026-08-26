"""Excel export — два формата: для менеджера (сводный склад) и для организатора (по наборам)."""
import io
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from ..models import Event, GiftSet

BRAND      = "2E4057"
BRAND_PALE = "EEF3FB"

def _side(style="thin", color="CCCCCC"):
    return Side(style=style, color=color)

def _border(color="CCCCCC"):
    s = _side(color=color)
    return Border(left=s, right=s, top=s, bottom=s)

BORDER_THIN  = _border()
BORDER_BRAND = _border(BRAND)

def _f(bold=False, size=10, color="000000"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def _c(ws, row, col, value="", bold=False, size=10, color="000000",
       bg=None, align="left", border=True, num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _f(bold=bold, size=size, color=color)
    c.alignment = Alignment(horizontal=align, vertical="center")
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if border:
        c.border = BORDER_THIN
    if num_fmt:
        c.number_format = num_fmt
    return c


_TYPE_ORDER = {"sample": 0, "pigment": 1, "consumable": 2, "certificate": 3}


def _get_set_count(gs: GiftSet, event: Event) -> int:
    if gs.place == "участник":
        return max(event.participants_count or 1, 1)
    if gs.place == "гран-при":
        return 1
    if gs.place == "розыгрыш":
        gm = getattr(event, "giveaway_mode", None) or "одинаковые"
        return 1 if gm == "разные" else max(event.giveaways_count or 1, 1)
    noms = event.nominations_data or []
    key = {"1": "place1", "2": "place2", "3": "place3"}.get(gs.place, "place1")
    for n in noms:
        if n.get("name") == gs.nomination_name:
            return max(n.get(key, 1), 1)
    return 1


def export_event_to_excel(event: Event, sets: list[GiftSet]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Список отгрузки"

    # Ширины колонок: # | Наименование | Линия | Цена | Кол-во | Итого
    for col, w in zip("ABCDEF", [4, 44, 20, 14, 12, 16]):
        ws.column_dimensions[col].width = w

    row = 1

    # ── Шапка мероприятия ────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:F{row}")
    t = ws.cell(row=row, column=1, value="FACE Pigments — список отгрузки")
    t.font = _f(bold=True, size=14, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=BRAND)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 30
    row += 1

    ws.merge_cells(f"A{row}:F{row}")
    n = ws.cell(row=row, column=1, value=event.name or "")
    n.font = _f(bold=True, size=12, color=BRAND)
    n.fill = PatternFill("solid", fgColor=BRAND_PALE)
    n.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 24
    row += 1

    def meta(label, value):
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = _f(bold=True, size=10, color=BRAND)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        ws.cell(row=row, column=2, value=value).font = _f(size=10)
        ws.row_dimensions[row].height = 17
        row += 1

    loc = event.country or ""
    if event.region and event.region != event.country:
        loc += f" / {event.region}"
    meta("Страна / регион:", loc)
    meta("Дата:", str(event.date or ""))
    meta("Склад:", event.warehouse or "")

    extras = []
    if getattr(event, "has_trade_booth", False):     extras.append("торговая точка")
    if getattr(event, "has_speaker_nonstop", False): extras.append("спикер нонстоп")
    if getattr(event, "has_speaker_stage", False):   extras.append("спикер на сцене")
    meta("Доп. условия:", ", ".join(extras) if extras else "—")

    row += 1  # отступ

    # ── Агрегация позиций ─────────────────────────────────────────────────
    # Ключ: (sku_type, sku_id, name, line) → {qty, price}
    agg: dict = defaultdict(lambda: {"qty": 0, "price": 0})

    for gs in sets:
        count = _get_set_count(gs, event)
        for item in gs.items or []:
            key = (
                item.get("sku_type", "consumable"),
                item.get("sku_id", 0),
                item.get("name", ""),
                item.get("line") or "",
            )
            agg[key]["qty"]   += item.get("qty", 1) * count
            agg[key]["price"]  = item.get("price", 0)  # цена единицы одинакова везде

    # Сортировка: сэмплы → пигменты (по линии, затем имени) → расходники → сертификаты
    def sort_key(entry):
        (sku_type, _, name, line) = entry[0]
        return (_TYPE_ORDER.get(sku_type, 9), line.lower(), name.lower())

    sorted_items = sorted(agg.items(), key=sort_key)

    # ── Заголовки таблицы ─────────────────────────────────────────────────
    col_labels = ["#", "Наименование", "Линия / категория", "Розн. цена", "Кол-во", "Итого"]
    col_aligns = ["center", "left", "left", "right", "center", "right"]
    for ci, (lbl, al) in enumerate(zip(col_labels, col_aligns), 1):
        c = ws.cell(row=row, column=ci, value=lbl)
        c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=BRAND)
        c.alignment = Alignment(horizontal=al, vertical="center")
        c.border = BORDER_THIN
    ws.row_dimensions[row].height = 20
    row += 1

    # ── Строки данных ─────────────────────────────────────────────────────
    STRIPE = "F5F8FC"
    grand_total = 0

    for idx, ((sku_type, sku_id, name, line), data) in enumerate(sorted_items, 1):
        qty   = data["qty"]
        price = data["price"]
        total = price * qty
        grand_total += total

        bg = STRIPE if idx % 2 == 0 else None

        _c(ws, row, 1, idx,   align="center", bg=bg)
        _c(ws, row, 2, name,  align="left",   bg=bg)
        _c(ws, row, 3, line,  align="left",   bg=bg, color="666666")
        _c(ws, row, 4, price, align="right",  bg=bg, num_fmt='#,##0 "₽"')
        _c(ws, row, 5, qty,   align="center", bg=bg)
        _c(ws, row, 6, total, align="right",  bg=bg, bold=True, num_fmt='#,##0 "₽"')
        ws.row_dimensions[row].height = 17
        row += 1

    # ── Итого ─────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    gl = ws.cell(row=row, column=1, value="ИТОГО К ОТГРУЗКЕ:")
    gl.font = _f(bold=True, size=12, color="FFFFFF")
    gl.fill = PatternFill("solid", fgColor=BRAND)
    gl.alignment = Alignment(horizontal="right", vertical="center")
    gl.border = BORDER_BRAND

    gv = ws.cell(row=row, column=6, value=grand_total)
    gv.font = _f(bold=True, size=12, color="FFFFFF")
    gv.fill = PatternFill("solid", fgColor=BRAND)
    gv.alignment = Alignment(horizontal="right", vertical="center")
    gv.number_format = '#,##0 "₽"'
    gv.border = BORDER_BRAND
    ws.row_dimensions[row].height = 28

    # Закрепить строку с заголовками таблицы
    ws.freeze_panes = f"A{row - len(sorted_items)}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─── Organizer export (no prices) ────────────────────────────────────────────

PLACE_LABELS_RU = {
    "1": "1 место", "2": "2 место", "3": "3 место",
    "гран-при": "Гран-при", "розыгрыш": "Розыгрыш", "участник": "Участник",
}
SKU_TYPE_RU = {
    "pigment": "Пигмент", "sample": "Мини-сэт",
    "consumable": "Расходник", "certificate": "Сертификат",
}


def export_event_organizer(event: Event, sets: list[GiftSet]) -> bytes:
    """
    Таблица для организатора мероприятия — без цен.
    Строки: Номинация | Место | Кол-во чел. | Тип | Позиция | Кол-во
    Ячейки Номинация/Место/Кол-во чел. объединяются по количеству позиций в наборе.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Для организатора"

    # Ширины: Номинация | Место | Чел. | Тип | Позиция | Кол-во
    for col, w in zip("ABCDEF", [32, 14, 10, 14, 44, 10]):
        ws.column_dimensions[get_column_letter(col if isinstance(col, int) else ord(col) - 64)].width = w
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 44
    ws.column_dimensions["F"].width = 10

    row = 1

    # Шапка
    ws.merge_cells(f"A{row}:F{row}")
    t = ws.cell(row=row, column=1, value="FACE Pigments — программа подарков")
    t.font = _f(bold=True, size=14, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=BRAND)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 30
    row += 1

    ws.merge_cells(f"A{row}:F{row}")
    n = ws.cell(row=row, column=1, value=event.name or "")
    n.font = _f(bold=True, size=12, color=BRAND)
    n.fill = PatternFill("solid", fgColor=BRAND_PALE)
    n.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22
    row += 1

    for label, value in [
        ("Дата:", str(event.date or "")),
        ("Место:", f"{event.country or ''} / {event.region or ''}".strip(" /")),
    ]:
        ws.cell(row=row, column=1, value=label).font = _f(bold=True, color=BRAND)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        ws.cell(row=row, column=2, value=value).font = _f()
        ws.row_dimensions[row].height = 16
        row += 1

    row += 1  # отступ

    # Заголовки таблицы
    headers = ["Номинация", "Место", "Чел.", "Тип", "Позиция", "Кол-во"]
    aligns  = ["left", "center", "center", "left", "left", "center"]
    for ci, (h, al) in enumerate(zip(headers, aligns), 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=BRAND)
        c.alignment = Alignment(horizontal=al, vertical="center")
        c.border = BORDER_THIN
    ws.row_dimensions[row].height = 20
    header_row = row
    row += 1

    STRIPE = "F5F8FC"

    # Сортировка наборов: гран-при → 1 → 2 → 3 → розыгрыш → участник
    place_order = {"гран-при": 0, "1": 1, "2": 2, "3": 3, "розыгрыш": 4, "участник": 5}

    def sort_items_for_display(items):
        order = {"sample": 0, "pigment": 1, "consumable": 2, "certificate": 3}
        return sorted(items, key=lambda i: (order.get(i.get("sku_type", "consumable"), 9), i.get("name", "")))

    sorted_sets = sorted(sets, key=lambda s: (place_order.get(s.place, 9), s.nomination_name or ""))

    for gs_idx, gs in enumerate(sorted_sets):
        count = _get_set_count(gs, event)
        items = sort_items_for_display(gs.items or [])
        n_items = len(items)
        if n_items == 0:
            continue

        bg = STRIPE if gs_idx % 2 == 0 else None
        start_row = row

        for item_idx, item in enumerate(items):
            name = item.get("name", "")
            line = item.get("line") or item.get("category") or ""
            if line:
                name = f"{name}  ({line})"
            qty = item.get("qty", 1)
            sku_type = SKU_TYPE_RU.get(item.get("sku_type", "consumable"), "")

            if item_idx == 0:
                _c(ws, row, 1, gs.nomination_name or "", bold=True, bg=bg)
                _c(ws, row, 2, PLACE_LABELS_RU.get(gs.place, gs.place), align="center", bg=bg)
                _c(ws, row, 3, count, align="center", bg=bg)
            else:
                _c(ws, row, 1, "", bg=bg)
                _c(ws, row, 2, "", bg=bg)
                _c(ws, row, 3, "", bg=bg)

            _c(ws, row, 4, sku_type, bg=bg, color="555555")
            _c(ws, row, 5, name, bg=bg)
            _c(ws, row, 6, qty, align="center", bg=bg)
            ws.row_dimensions[row].height = 16
            row += 1

        # Merge nomination/place/count cells across all item rows
        if n_items > 1:
            for col in (1, 2, 3):
                ws.merge_cells(start_row=start_row, start_column=col,
                               end_row=row - 1,    end_column=col)
                merged_cell = ws.cell(row=start_row, column=col)
                merged_cell.alignment = Alignment(horizontal="center" if col > 1 else "left",
                                                  vertical="top", wrap_text=True)

    ws.freeze_panes = f"A{header_row + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
